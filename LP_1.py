

import pulp
import openpyxl


class Employee ():
    def __init__(self,eid, pay, lowhours, highhours, availability):
                self.eid = eid
                self.pay = pay
                self.lowhours = lowhours
                self.highhours = highhours
                self.availability = availability
                
     

def readexcelfile(path, employees, demandRoles):
    
    wb = openpyxl.load_workbook(filename = path)
    sheetnames = wb.sheetnames
    wsheet =wb[sheetnames[0]]
    print (wsheet.max_column, wsheet.max_row)
    for j in range (1,wsheet.max_column):
        eid = wsheet[1][j].value
        pay = float(wsheet[2][j].value)
        #print (pay)
        lowhours = int(wsheet[3][j].value)
        highhours = int(wsheet[4][j].value)
        roles = [int(wsheet[5][j].value),int(wsheet[6][j].value),int(wsheet[7][j].value),int(wsheet[8][j].value),int(wsheet[9][j].value)]
        availability_temp=[]
        unavailability_temp=[]
        for i in range(0,7):
            #print(i)
            temp=[]
            un_temp=[]
            for k in range (11+i*10, 21+i*10):
                #print(j)
                temp.append(int(wsheet[k][j].value))
                un_temp.append(0)
            availability_temp.append(temp)
            unavailability_temp.append(un_temp)
        availability=[]
        for i in range(len(roles)):
            if(roles[i]==1):
                availability.append(availability_temp)
            else:
                availability.append(unavailability_temp)
        employees.append(Employee(eid, pay, lowhours, highhours, availability))
    wsheet =wb[sheetnames[1]]
    for k in range(0,5):
        tempDemand=[]
        for j in range(0,7):
            temptempDemand=[]
            for i in range (2+j*10,12+j*10):
                temptempDemand.append(int(wsheet[i][k+1].value))
            tempDemand.append(temptempDemand)
        demandRoles.append(tempDemand)

         
        

path ="Project3data.xlsx"



employees=[]
demandRoles=[]

readexcelfile(path, employees, demandRoles)
'''
print('The demands for cashiers by the company are ')
for i in range (0,len(rcashier)):
    print (rcashier[i], rstocking[i])

print('Information on the employees ')
for i in range (0,len(employees)):
    print (employees[i].eid, employees[i].pay)
    for j in range (0,len(employees[i].availability)):
        if employees[i].availability[j]==0:
            print (j, 'Not available at that time')
        if employees[i].availability[j]==1:
            print (j, 'Available at that time')
'''
numEmployees=len(employees)
numDays=7
numHours=10
numRoles=5
print(numEmployees)

#weight of objecctive set
alpha=0.8

#Define LP Problem
my_lp_problem = pulp.LpProblem("My LP Problem", pulp.LpMinimize)

#LP Variable sets
lpWorkedHours= pulp.LpVariable.dicts("workedHours", ((i, r , d, h) for i in range(numEmployees) for r in range(numRoles) for d in range(numDays) for h in range(numHours)), lowBound=0, upBound=1, cat = 'Integer')

#Objective Function that needs to be minimized
my_lp_problem += pulp.lpSum([lpWorkedHours[i,r,d,h]*employees[i].pay for i in range(numEmployees) for r in range(numRoles) for d in range(numDays) for h in range(numHours)])

#Availability Constraint
for i in range(numEmployees):
    for r in range(numRoles):
        for d in range(numDays):
            for h in range(numHours):
                my_lp_problem += lpWorkedHours[i,r,d,h]<=employees[i].availability[r][d][h]


#Employee One role in one hour constraint
for i in range(numEmployees):
    for d in range(numDays):
        for h in range(numHours):
            my_lp_problem+=pulp.lpSum([lpWorkedHours[i,r,d,h]for r in range(numRoles)])<=1

#Employee working desire constraint max
for i in range(numEmployees):
    my_lp_problem+=pulp.lpSum([lpWorkedHours[i,r,d,h]for r in range(numRoles) for d in range(numDays) for h in range(numHours)])<=employees[i].highhours

#Employee working desire constraint min
for i in range(numEmployees):
    my_lp_problem+=pulp.lpSum([lpWorkedHours[i,r,d,h]for r in range(numRoles) for d in range(numDays) for h in range(numHours)])>=employees[i].lowhours

#Wrokforce Requirement Constraint
for r in range(numRoles):
    for d in range(numDays):
        for h in range(numHours):
            my_lp_problem+=pulp.lpSum([lpWorkedHours[i,r,d,h] for i in range(numEmployees)])>=demandRoles[r][d][h]


#Prinitning my solution
#print (my_lp_problem)
status=my_lp_problem.solve()
if pulp.LpStatus[my_lp_problem.status] =='Infeasible':
    print ('INFEASIBLE ****************************************')
else:
    #This is various ways to display the solution
 
    for variable in my_lp_problem.variables():
        print ("{} = {}".format(variable.name, variable.varValue))
    print ('\nFeasible\n\nThe objective value is ', pulp.value(my_lp_problem.objective),'\n')

#To help you write out the solution.

'''
wbr = openpyxl.Workbook()

#Sheets start at 0 

wbr.create_sheet(index = 0, title = "solution")
sheetnames = wbr.sheetnames



wsheet = wbr[sheetnames[0]]

#Writing to individual cells
# Cells start at 1, 1

#set column widths





row =1 
wsheet.cell(row,1).value ='Employee costs'
wsheet.cell(row,2).value = my_lp_problem.objective
row+=2

for i in range (0, len(employees)):
    wsheet.cell(row , i+2).value = employees[i].eid


days=0
hours=10
for i in range (0, numDays*numHours-1):
    row+=1
    wsheet.cell(row , 1).value = str(days)+','+str(hours)

    hours+=1
    if hours==21:
        hours=10
        days+=1
    

wbr.save('Project3Solution.xlsx')'''